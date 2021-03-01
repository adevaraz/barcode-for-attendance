import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def attendance(kd_mk, nim, nama):
    presence = datetime.now()
    date = presence.strftime("%d/%m/%Y")
    time = presence.strftime("%H:%M:%S")

    sheet = "MK_" + str(kd_mk)

    status = 'hadir'

    df = pd.DataFrame([[date, time, nim, nama, status]], 
                        columns=['tanggal', 'waktu_presensi', 'NIM', 'Nama', 'status'])
    
    book = load_workbook('./data/presensi.xlsx')
    writer = pd.ExcelWriter('./data/presensi.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
 
    if sheet in book.sheetnames:
        current_sheet = book[sheet]
        row_count = current_sheet.max_row
        try:
            df.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row_count)
            print("Presensi berhasil.")
        except:
            print("Gagal menyimpan presensi!")
    else:
        try:
            df.to_excel(writer, sheet_name=sheet, index=False)
            print("Presensi berhasil.")
        except:
            print("Gagal menyimpan presensi!")

    writer.close()
    book.close()

attendance(3, '181524002', 'Alvira')
    


