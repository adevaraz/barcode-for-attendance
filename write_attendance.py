import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from compute_attendance_time import compute_attendance_time
import constants


def attendance(lecture_id, student_id, name):
    presence = datetime.now()
    date = presence.strftime("%d/%m/%Y")
    time = presence.strftime("%H:%M:%S")

    sheet = "MK_" + str(lecture_id)

    status = compute_attendance_time(datetime.now(), student_id)

    df = pd.DataFrame([[date, time, student_id, name, status]], 
                        columns=['tanggal', 'waktu_presensi', 'NIM', 'Nama', 'status'])
    
    book = load_workbook(constants.PRESENSI_LOC)
    writer = pd.ExcelWriter(constants.PRESENSI_LOC, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
 
    if sheet in book.sheetnames:
        current_sheet = book[sheet]
        row_count = current_sheet.max_row
        try:
            df.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=row_count)
            print("Presensi berhasil.")
        except:
            print("Gagal menyimpan presensi!")
    else:
        try:
            df.to_excel(writer, sheet_name=sheet, index=False)
            print("Presensi berhasil.")
        except:
            print("Gagal menyimpan presensi!")

    writer.close()
    book.close()

attendance(3, '181524007', 'Evan')
    


